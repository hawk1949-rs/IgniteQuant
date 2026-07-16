#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""回测存档：可选接入，结束后在桌面生成 Excel 对账单。

用法（策略内按需接入，非强制）::

    from common.backtest_archive import BacktestArchive

    sim = TqSim(init_balance=20_000_000)
    archive = BacktestArchive(
        strategy_name="Falcon v2",
        symbol="KQ.m@SHFE.au",
        backtest_start=START_DT,
        backtest_end=END_DT,
        init_balance=20_000_000,
        sim_account=sim,  # 用于汇总累计手续费/平仓盈亏
    )
    api = TqApi(sim, backtest=..., auth=...)
    # 下单前打标信号强度；每轮 wait_update 后 poll
    archive.tag_next(signal_strength=2, regime="TREND_UP", note="开多")
    target_pos.set_target_volume(3)
    archive.poll(api)
    # 回测结束
    path = archive.save(api)
"""

from __future__ import annotations

import datetime as dt
import re
from collections import defaultdict, deque
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, DefaultDict, Deque, Dict, List, Mapping, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

Number = Union[int, float]
SignalStrength = Union[int, float, str]


def _to_float(value: Any, default: float = 0.0) -> float:
    """把账户/成交字段转成可写入 Excel 的 Python float（nan→default）。"""
    if value is None:
        return default
    try:
        f = float(value)
    except (TypeError, ValueError):
        return default
    if f != f:  # NaN
        return default
    return f


def _to_excel_number(value: Any) -> Union[int, float]:
    f = _to_float(value, 0.0)
    if abs(f - round(f)) < 1e-9:
        return int(round(f))
    return round(f, 4)


def desktop_dir() -> Path:
    """解析本机桌面目录（兼容 Desktop / 桌面 / OneDrive）。"""
    home = Path.home()
    candidates = (
        home / "Desktop",
        home / "桌面",
        home / "OneDrive" / "Desktop",
        home / "OneDrive" / "桌面",
    )
    for path in candidates:
        if path.is_dir():
            return path
    fallback = home / "Desktop"
    fallback.mkdir(parents=True, exist_ok=True)
    return fallback


def _safe_filename(text: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\s]+', "_", text.strip())
    return cleaned.strip("._") or "backtest"


def _fmt_dt(value: Optional[dt.datetime]) -> str:
    if value is None:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _trade_datetime(trade: Any) -> Optional[dt.datetime]:
    raw = getattr(trade, "trade_date_time", None)
    if raw is None:
        raw = getattr(trade, "datetime", None)
    if raw is None:
        return None
    try:
        ns = int(raw)
    except (TypeError, ValueError):
        return None
    # 天勤成交时间多为纳秒；若已是秒级则直接用
    if ns > 10_000_000_000_000:
        seconds = ns / 1_000_000_000
    else:
        seconds = float(ns)
    try:
        return dt.datetime.fromtimestamp(seconds)
    except (OSError, OverflowError, ValueError):
        return None


def _map_direction(raw: str) -> str:
    key = (raw or "").upper()
    return {"BUY": "买", "SELL": "卖"}.get(key, raw or "")


def _map_offset(raw: str) -> str:
    key = (raw or "").upper()
    return {
        "OPEN": "开",
        "CLOSE": "平",
        "CLOSETODAY": "平今",
        "CLOSEYESTERDAY": "平昨",
    }.get(key, raw or "")


@dataclass
class SignalTag:
    strength: SignalStrength
    regime: str = ""
    note: str = ""
    parts: str = ""


@dataclass
class TradeRow:
    trade_id: str
    trade_time: Optional[dt.datetime]
    symbol: str
    direction: str
    offset: str
    volume: Number
    price: Number
    commission: Number
    close_profit: Number
    signal_strength: SignalStrength
    regime: str = ""
    note: str = ""
    parts: str = ""


@dataclass
class BacktestArchive:
    """可插拔回测存档器：收集成交并导出桌面 Excel。"""

    strategy_name: str
    symbol: str
    launched_at: dt.datetime = field(default_factory=dt.datetime.now)
    save_dir: Optional[Path] = None
    backtest_start: Optional[dt.date] = None
    backtest_end: Optional[dt.date] = None
    init_balance: Optional[float] = None
    # 传入 TqSim 实例时可读取 trade_log / tqsdk_stat，汇总累计手续费与平仓盈亏
    sim_account: Any = None

    _tags: deque[SignalTag] = field(default_factory=deque, init=False, repr=False)
    _last_tag: Optional[SignalTag] = field(default=None, init=False, repr=False)
    _seen_trade_ids: set[str] = field(default_factory=set, init=False, repr=False)
    _rows: list[TradeRow] = field(default_factory=list, init=False, repr=False)
    _account_snapshot: dict[str, Any] = field(default_factory=dict, init=False, repr=False)
    _saved_path: Optional[Path] = field(default=None, init=False, repr=False)
    _multiples: Dict[str, float] = field(default_factory=dict, init=False, repr=False)

    def tag_next(
        self,
        signal_strength: SignalStrength,
        *,
        regime: str = "",
        note: str = "",
        parts: str = "",
    ) -> None:
        """为随后产生的成交打上信号强度（FIFO；一笔调仓通常对应一条成交）。"""
        tag = SignalTag(
            strength=signal_strength,
            regime=regime,
            note=note,
            parts=parts,
        )
        self._tags.append(tag)
        self._last_tag = tag

    def record_manual(
        self,
        *,
        trade_time: Optional[dt.datetime],
        symbol: str,
        direction: str,
        offset: str,
        volume: Number,
        price: Number,
        signal_strength: SignalStrength,
        commission: Number = 0,
        close_profit: Number = 0,
        regime: str = "",
        note: str = "",
        parts: str = "",
        trade_id: str = "",
    ) -> None:
        """不依赖 api.get_trade 时，由策略直接写入一笔明细。"""
        tid = trade_id or f"manual-{len(self._rows) + 1}"
        if tid in self._seen_trade_ids:
            return
        self._seen_trade_ids.add(tid)
        self._rows.append(
            TradeRow(
                trade_id=tid,
                trade_time=trade_time,
                symbol=symbol,
                direction=_map_direction(direction) if direction.isascii() else direction,
                offset=_map_offset(offset) if offset.isascii() else offset,
                volume=volume,
                price=price,
                commission=commission,
                close_profit=close_profit,
                signal_strength=signal_strength,
                regime=regime,
                note=note,
                parts=parts,
            )
        )

    def poll(self, api: Any) -> int:
        """同步 ``api.get_trade()`` 中尚未归档的成交，返回新增条数。"""
        try:
            trades = api.get_trade()
        except Exception:
            return 0
        if trades is None:
            return 0

        added = 0
        items: list[tuple[str, Any]]
        if isinstance(trades, Mapping):
            items = list(trades.items())
        else:
            # 兼容类 dict / 可迭代
            try:
                items = [(str(t.trade_id), t) for t in trades]
            except TypeError:
                return 0

        # 按成交时间排序，保证 tag FIFO 与时间线一致
        def sort_key(item: tuple[str, Any]) -> tuple[int, str]:
            t = item[1]
            raw = getattr(t, "trade_date_time", None) or getattr(t, "datetime", 0) or 0
            try:
                return (int(raw), item[0])
            except (TypeError, ValueError):
                return (0, item[0])

        for trade_id, trade in sorted(items, key=sort_key):
            tid = str(trade_id)
            if tid in self._seen_trade_ids:
                continue
            self._seen_trade_ids.add(tid)

            tag = self._tags.popleft() if self._tags else self._last_tag
            strength: SignalStrength = tag.strength if tag else ""
            regime = tag.regime if tag else ""
            note = tag.note if tag else ""
            parts = tag.parts if tag else ""

            symbol = (
                getattr(trade, "symbol", None)
                or getattr(trade, "instrument_id", None)
                or ""
            )
            exchange = getattr(trade, "exchange_id", "") or ""
            if symbol and exchange and "." not in str(symbol):
                symbol = f"{exchange}.{symbol}"
            symbol = str(symbol)
            self._cache_volume_multiple(api, symbol)

            # Trade 对象无平仓盈亏字段；先占位，save 前用 FIFO 回算
            raw_cp = getattr(trade, "close_profit", None)
            if raw_cp is None:
                raw_cp = getattr(trade, "profit", None)

            self._rows.append(
                TradeRow(
                    trade_id=tid,
                    trade_time=_trade_datetime(trade),
                    symbol=symbol,
                    direction=_map_direction(str(getattr(trade, "direction", "") or "")),
                    offset=_map_offset(str(getattr(trade, "offset", "") or "")),
                    volume=_num(getattr(trade, "volume", 0)),
                    price=_num(getattr(trade, "price", 0)),
                    commission=_num(getattr(trade, "commission", 0)),
                    close_profit=_num(raw_cp) if raw_cp is not None else 0,
                    signal_strength=strength,
                    regime=regime,
                    note=note,
                    parts=parts,
                )
            )
            added += 1
        return added

    def _cache_volume_multiple(self, api: Any, symbol: str) -> float:
        if symbol in self._multiples:
            return self._multiples[symbol]
        multiple = 0.0
        if api is not None and symbol:
            try:
                quote = api.get_quote(symbol)
                multiple = _to_float(getattr(quote, "volume_multiple", 0))
            except Exception:
                multiple = 0.0
        if multiple <= 0:
            # 常见品种兜底：沪金 1000 克/手
            low = symbol.lower()
            if "au" in low and "autd" not in low:
                multiple = 1000.0
            else:
                multiple = 1.0
        self._multiples[symbol] = multiple
        return multiple

    def recompute_close_profits(self) -> None:
        """按开平 FIFO 回算每笔成交的平仓盈亏（天勤 Trade 不含该字段）。"""
        books: DefaultDict[str, Dict[str, Deque[List[float]]]] = defaultdict(
            lambda: {"long": deque(), "short": deque()}
        )

        ordered = sorted(
            self._rows,
            key=lambda r: (
                r.trade_time or dt.datetime.min,
                r.trade_id,
            ),
        )
        for row in ordered:
            multiple = self._multiples.get(row.symbol) or self._guess_multiple(row.symbol)
            self._multiples[row.symbol] = multiple
            offset_raw = row.offset or ""
            offset_u = offset_raw.upper()
            is_open = offset_raw == "开" or offset_u == "OPEN"
            is_close = (
                offset_raw in ("平", "平今", "平昨")
                or offset_u in ("CLOSE", "CLOSETODAY", "CLOSEYESTERDAY")
                or offset_raw.startswith("平")
            )
            if row.direction in ("买", "卖"):
                is_buy = row.direction == "买"
            else:
                is_buy = (row.direction or "").upper() == "BUY"

            vol = int(_to_float(row.volume))
            price = _to_float(row.price)
            if vol <= 0:
                row.close_profit = 0
                continue

            if is_open or not is_close:
                side = "long" if is_buy else "short"
                books[row.symbol][side].append([price, float(vol)])
                row.close_profit = 0
                continue

            # 平仓：卖平多 / 买平空
            side = "long" if not is_buy else "short"
            book = books[row.symbol][side]
            pnl = 0.0
            left = float(vol)
            while left > 0 and book:
                open_price, open_vol = book[0]
                take = min(open_vol, left)
                if side == "long":
                    pnl += (price - open_price) * take * multiple
                else:
                    pnl += (open_price - price) * take * multiple
                open_vol -= take
                left -= take
                if open_vol <= 1e-9:
                    book.popleft()
                else:
                    book[0][1] = open_vol
            row.close_profit = _to_excel_number(pnl)

    @staticmethod
    def _guess_multiple(symbol: str) -> float:
        low = (symbol or "").lower()
        if "au" in low and "autd" not in low:
            return 1000.0
        return 1.0

    def capture_account(self, api: Any = None) -> None:
        """汇总概要字段。

        注意：``api.get_account()`` 里的平仓盈亏/手续费是**当日截面**，回测末日常为 0。
        累计值优先从 ``TqSim.trade_log`` / ``tqsdk_stat`` 与成交明细汇总。
        """
        snap: dict[str, Any] = {
            "trade_count": len(self._rows),
            "init_balance": _to_excel_number(self.init_balance or 0),
            "balance": 0,
            "available": 0,
            "margin": 0,
            "margin_peak": 0,
            "margin_avg": 0,
            "float_profit": 0,
            "position_profit": 0,
            "close_profit": 0,
            "commission": 0,
            "risk_ratio": 0.0,
        }

        # 1) 实时账户截面（期末持仓相关字段以这里为准）
        if api is not None:
            try:
                acc = api.get_account()
                for key in (
                    "balance",
                    "available",
                    "margin",
                    "float_profit",
                    "position_profit",
                    "close_profit",
                    "commission",
                    "risk_ratio",
                ):
                    if hasattr(acc, key):
                        snap[key] = _to_excel_number(getattr(acc, key))
                pre = getattr(acc, "pre_balance", None)
                if self.init_balance is None and pre is not None:
                    snap["init_balance"] = _to_excel_number(pre)
            except Exception:
                pass

        sim = self.sim_account
        if sim is None and api is not None:
            sim = getattr(api, "_account", None)

        # 2) 按日 trade_log 累加：平仓盈亏、手续费；并用末日账户覆盖权益类字段
        trade_log = getattr(sim, "trade_log", None) if sim is not None else None
        if isinstance(trade_log, Mapping) and trade_log:
            sum_close = 0.0
            sum_commission = 0.0
            margins: List[float] = []
            last_acc: dict[str, Any] = {}
            for day in sorted(trade_log.keys()):
                day_acc = trade_log[day].get("account") or {}
                if isinstance(day_acc, Mapping):
                    last_acc = dict(day_acc)
                    sum_close += _to_float(day_acc.get("close_profit"))
                    sum_commission += _to_float(day_acc.get("commission"))
                    margins.append(_to_float(day_acc.get("margin")))
            for key in (
                "balance",
                "available",
                "margin",
                "float_profit",
                "position_profit",
                "risk_ratio",
            ):
                if key in last_acc:
                    snap[key] = _to_excel_number(last_acc[key])
            if "pre_balance" in (trade_log[sorted(trade_log.keys())[0]].get("account") or {}):
                first_acc = trade_log[sorted(trade_log.keys())[0]].get("account") or {}
                if self.init_balance is None:
                    snap["init_balance"] = _to_excel_number(first_acc.get("pre_balance"))
            snap["close_profit"] = _to_excel_number(sum_close)
            snap["commission"] = _to_excel_number(sum_commission)
            if margins:
                snap["margin_peak"] = _to_excel_number(max(margins))
                occupied = [m for m in margins if m > 0]
                snap["margin_avg"] = _to_excel_number(
                    (sum(occupied) / len(occupied)) if occupied else 0.0
                )

        # 3) 成交明细汇总手续费 / 平仓盈亏（trade_log 缺失时的兜底）
        row_commission = sum(_to_float(r.commission) for r in self._rows)
        row_close = sum(_to_float(r.close_profit) for r in self._rows)
        if _to_float(snap["commission"]) <= 0 and row_commission > 0:
            snap["commission"] = _to_excel_number(row_commission)
        if abs(_to_float(snap["close_profit"])) < 1e-9 and abs(row_close) > 1e-9:
            snap["close_profit"] = _to_excel_number(row_close)

        # 4) tqsdk_stat 覆盖期末权益 / 初始资金 / 总手续费
        stat = getattr(sim, "tqsdk_stat", None) if sim is not None else None
        if isinstance(stat, Mapping) and stat:
            if "init_balance" in stat:
                snap["init_balance"] = _to_excel_number(stat["init_balance"])
            if "end_balance" in stat:
                snap["balance"] = _to_excel_number(stat["end_balance"])
            elif "balance" in stat:
                snap["balance"] = _to_excel_number(stat["balance"])
            if "commission" in stat:
                snap["commission"] = _to_excel_number(stat["commission"])
            snap["tqsdk_stat"] = {
                k: stat[k]
                for k in (
                    "ror",
                    "annual_yield",
                    "max_drawdown",
                    "winning_rate",
                    "profit_loss_ratio",
                    "sharpe_ratio",
                )
                if k in stat
            }

        init_bal = _to_float(snap["init_balance"])
        end_bal = _to_float(snap["balance"])
        snap["total_pnl"] = _to_excel_number(end_bal - init_bal)
        snap["trade_count"] = len(self._rows)
        self._account_snapshot = snap

    def default_path(self) -> Path:
        """默认导出路径（桌面 / 自定义 save_dir）。"""
        base = Path(self.save_dir) if self.save_dir else desktop_dir()
        stamp = self.launched_at.strftime("%Y%m%d_%H%M%S")
        name = (
            f"回测_{_safe_filename(self.strategy_name)}_"
            f"{_safe_filename(self.symbol)}_{stamp}.xlsx"
        )
        return base / name

    def save(self, api: Any = None, path: Optional[Path] = None) -> Path:
        """写出 Excel；默认桌面。若传入 api 会先 poll；始终汇总概要字段。"""
        if api is not None:
            self.poll(api)
            # 补齐合约乘数，供 FIFO 回算平仓盈亏
            for row in self._rows:
                if row.symbol and row.symbol not in self._multiples:
                    self._cache_volume_multiple(api, row.symbol)
        self.recompute_close_profits()
        self.capture_account(api)

        out = Path(path) if path is not None else self.default_path()
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        self._write_summary_sheet(wb.active)
        self._write_trades_sheet(wb.create_sheet("成交明细", 1))
        wb.save(out)
        self._saved_path = out
        return out

    @property
    def saved_path(self) -> Optional[Path]:
        return self._saved_path

    @property
    def trade_count(self) -> int:
        return len(self._rows)

    def _write_summary_sheet(self, ws: Any) -> None:
        ws.title = "回测概要"
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        label_font = Font(bold=True, color="000000")
        value_font = Font(color="000000")
        snap = self._account_snapshot

        ws["A1"] = "回测存档概要"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws.merge_cells("A1:B1")

        ws["A2"] = "项目"
        ws["B2"] = "数值"
        for col in (1, 2):
            cell = ws.cell(2, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        risk = _to_float(snap.get("risk_ratio"))
        rows: list[tuple[str, Any]] = [
            ("策略名称", self.strategy_name),
            ("品种/合约", self.symbol),
            ("回测启动时间", _fmt_dt(self.launched_at)),
            (
                "回测区间",
                f"{self.backtest_start or '-'} ~ {self.backtest_end or '-'}",
            ),
            ("导出时间", _fmt_dt(dt.datetime.now())),
            ("成交笔数", int(snap.get("trade_count", len(self._rows)) or 0)),
            ("初始资金", _to_excel_number(snap.get("init_balance", 0))),
            ("账户权益", _to_excel_number(snap.get("balance", 0))),
            ("可用资金", _to_excel_number(snap.get("available", 0))),
            ("期末保证金", _to_excel_number(snap.get("margin", 0))),
            ("峰值保证金", _to_excel_number(snap.get("margin_peak", 0))),
            ("占用日均保证金", _to_excel_number(snap.get("margin_avg", 0))),
            ("浮动盈亏", _to_excel_number(snap.get("float_profit", 0))),
            ("持仓盈亏", _to_excel_number(snap.get("position_profit", 0))),
            ("平仓盈亏", _to_excel_number(snap.get("close_profit", 0))),
            ("手续费", _to_excel_number(snap.get("commission", 0))),
            ("累计盈亏", _to_excel_number(snap.get("total_pnl", 0))),
            ("风险度", f"{risk * 100:.4f}%"),
        ]

        stat = snap.get("tqsdk_stat") if isinstance(snap.get("tqsdk_stat"), Mapping) else None
        if stat:
            if "ror" in stat:
                rows.append(("收益率", f"{_to_float(stat['ror']) * 100:.4f}%"))
            if "annual_yield" in stat:
                rows.append(("年化收益率", f"{_to_float(stat['annual_yield']) * 100:.4f}%"))
            if "max_drawdown" in stat:
                rows.append(("最大回撤", f"{_to_float(stat['max_drawdown']) * 100:.4f}%"))
            if "winning_rate" in stat:
                rows.append(("胜率", f"{_to_float(stat['winning_rate']) * 100:.4f}%"))
            if "sharpe_ratio" in stat:
                rows.append(("年化夏普", round(_to_float(stat["sharpe_ratio"]), 4)))

        for i, (label, value) in enumerate(rows, start=3):
            c1 = ws.cell(i, 1, label)
            c1.font = label_font
            c2 = ws.cell(i, 2, value)
            c2.font = value_font
            if isinstance(value, float):
                c2.number_format = "#,##0.00"
            elif isinstance(value, int) and abs(value) >= 1000:
                c2.number_format = "#,##0"

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 42

    def _write_trades_sheet(self, ws: Any) -> None:
        headers = [
            "序号",
            "成交时间",
            "合约",
            "买卖",
            "开平",
            "手数",
            "成交价",
            "手续费",
            "平仓盈亏",
            "信号强度",
            "行情状态",
            "信号分项",
            "备注",
            "成交编号",
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(1, col, title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for idx, row in enumerate(self._rows, start=1):
            values = [
                idx,
                _fmt_dt(row.trade_time),
                row.symbol,
                row.direction,
                row.offset,
                row.volume,
                row.price,
                row.commission,
                row.close_profit,
                row.signal_strength,
                row.regime,
                row.parts,
                row.note,
                row.trade_id,
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(idx + 1, col, value)

        widths = [6, 20, 16, 6, 8, 8, 12, 10, 12, 10, 12, 28, 24, 36]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(self._rows) + 1, 1)}"


def _num(value: Any) -> Number:
    if value is None:
        return 0
    try:
        f = float(value)
    except (TypeError, ValueError):
        return 0
    if f == int(f):
        return int(f)
    return f
