"""Overseas (外盘) instrument catalog + domestic pairing for research / cockpit."""

from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path


@dataclass(frozen=True)
class OverseasInstrumentSpec:
    id: str
    name: str
    signal_symbol: str  # archive / Yahoo symbol, e.g. GC=F
    exchange: str
    yahoo_symbol: str
    eastmoney_secid: str
    display_symbol: str
    currency: str = "USD"
    multiplier: float = 100.0
    tick_size: float = 0.1
    note: str = ""


@dataclass(frozen=True)
class OverseasPair:
    domestic_id: str
    overseas_id: str


# Overseas gold/silver: live signal clock prefers London spot (东财 122.XAU/XAG).
# COMEX continuous (GC00Y) trades at a futures premium and must not be labeled as XAUUSD.
OVERSEAS_INSTRUMENTS: dict[str, OverseasInstrumentSpec] = {
    "gc": OverseasInstrumentSpec(
        id="gc",
        name="伦敦金（现货黄金）",
        signal_symbol="XAUUSD",
        exchange="OTC",
        yahoo_symbol="XAUUSD=X",
        eastmoney_secid="122.XAU",
        display_symbol="XAUUSD",
        multiplier=100.0,
        tick_size=0.01,
        note="伦敦金现货（东财 122.XAU）；对照沪金 au。非 COMEX 期货 GC00Y。",
    ),
    "si": OverseasInstrumentSpec(
        id="si",
        name="伦敦银（现货白银）",
        signal_symbol="XAGUSD",
        exchange="OTC",
        yahoo_symbol="XAGUSD=X",
        eastmoney_secid="122.XAG",
        display_symbol="XAGUSD",
        multiplier=5000.0,
        tick_size=0.001,
        note="伦敦银现货（东财 122.XAG）；对照沪银 ag。非 COMEX 期货 SI00Y。",
    ),
    "hg": OverseasInstrumentSpec(
        id="hg",
        name="COMEX铜",
        signal_symbol="HG=F",
        exchange="COMEX",
        yahoo_symbol="HG=F",
        eastmoney_secid="101.HG00Y",
        display_symbol="HG",
        multiplier=25000.0,
        tick_size=0.0005,
        note="COMEX Copper continuous",
    ),
    "cl": OverseasInstrumentSpec(
        id="cl",
        name="NYMEX原油",
        signal_symbol="CL=F",
        exchange="NYMEX",
        yahoo_symbol="CL=F",
        eastmoney_secid="102.CL00Y",
        display_symbol="CL",
        multiplier=1000.0,
        tick_size=0.01,
        note="NYMEX WTI continuous",
    ),
}


# Domestic cockpit symbol_id → overseas id (for dual-chart / paired research).
OVERSEAS_PAIRS: dict[str, OverseasPair] = {
    "au": OverseasPair(domestic_id="au", overseas_id="gc"),
    "ag": OverseasPair(domestic_id="ag", overseas_id="si"),
}

# Phase-1 Strategy Lab: only gold/silver are wired for overseas backtest + cache.
LAB_OVERSEAS_WIRED: frozenset[str] = frozenset({"au", "ag"})

_ROOT = Path(__file__).resolve().parents[3]
_EXCEL_PATH = _ROOT / "data" / "内外盘品种对照.xlsx"


def overseas_by_id(product_id: str) -> OverseasInstrumentSpec:
    key = product_id.strip().lower()
    if key not in OVERSEAS_INSTRUMENTS:
        raise KeyError(f"unsupported overseas id: {product_id}")
    return OVERSEAS_INSTRUMENTS[key]


@lru_cache(maxsize=1)
def domestic_codes_from_excel(path: str | None = None) -> frozenset[str]:
    """Parse futures product codes (AU/AG/…) from the对照表; empty on failure."""
    xlsx = Path(path) if path else _EXCEL_PATH
    if not xlsx.is_file():
        return frozenset()
    try:
        import openpyxl
    except ImportError:
        return frozenset()
    try:
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return frozenset()
    if not rows:
        return frozenset()
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    code_idx = None
    for i, h in enumerate(header):
        if "代码" in h or h.upper() in {"CODE", "SYMBOL", "PRODUCT"}:
            code_idx = i
            break
    if code_idx is None:
        # Fallback: third column in the shipped workbook.
        code_idx = 2 if len(header) > 2 else 0
    codes: set[str] = set()
    for row in rows[1:]:
        if not row or code_idx >= len(row):
            continue
        raw = row[code_idx]
        if raw is None:
            continue
        text = str(raw).strip().upper()
        if not text or not text.replace("_", "").isalnum():
            continue
        # RU / AU / AG / SC …
        if 1 <= len(text) <= 4 and text.isalpha():
            codes.add(text.lower())
    return frozenset(codes)


def lab_overseas_supported(domestic_id: str) -> bool:
    """True when Strategy Lab may enable overseas pricing for this domestic id."""
    key = domestic_id.strip().lower()
    if key not in LAB_OVERSEAS_WIRED:
        return False
    if key not in OVERSEAS_PAIRS:
        return False
    excel = domestic_codes_from_excel()
    if excel and key not in excel:
        return False
    return True


def overseas_pair_for_domestic(domestic_id: str) -> dict[str, str] | None:
    """Research helper: domestic id → overseas instrument fields."""
    pair = OVERSEAS_PAIRS.get(domestic_id.strip().lower())
    if pair is None:
        return None
    spec = OVERSEAS_INSTRUMENTS[pair.overseas_id]
    return {
        "id": spec.id,
        "name": spec.name,
        "display_symbol": spec.display_symbol,
        "yahoo_symbol": spec.yahoo_symbol,
        "eastmoney_secid": spec.eastmoney_secid,
        "signal_symbol": spec.signal_symbol,
        "exchange": spec.exchange,
        "note": spec.note,
    }


def lab_overseas_pair_payload(domestic_id: str) -> dict[str, str] | None:
    """Compact pair for /api/catalog (Strategy Lab toggle)."""
    if not lab_overseas_supported(domestic_id):
        return None
    pair = overseas_pair_for_domestic(domestic_id)
    if not pair:
        return None
    return {
        "id": pair["id"],
        "name": pair["name"],
        "display_symbol": pair["display_symbol"],
    }


def cockpit_overseas_pair(domestic_id: str) -> dict[str, str] | None:
    """Catalog payload matching legacy OVERSEAS_PAIRS fields."""
    pair = OVERSEAS_PAIRS.get(domestic_id.strip().lower())
    if pair is None:
        return None
    spec = OVERSEAS_INSTRUMENTS[pair.overseas_id]
    legacy = {
        "gc": {
            "id": "xauusd",
            "name": "伦敦金（现货黄金）",
            "display_symbol": "XAUUSD",
            "note": (
                "外盘信号时钟：伦敦金现货（东财 122.XAU / XAUUSD）；"
                "不是 COMEX 期货 GC00Y（期货相对现货常有升水）。"
            ),
        },
        "si": {
            "id": "xagusd",
            "name": "伦敦银（现货白银）",
            "display_symbol": "XAGUSD",
            "note": (
                "外盘信号时钟：伦敦银现货（东财 122.XAG / XAGUSD）；"
                "不是 COMEX 期货 SI00Y。"
            ),
        },
    }.get(spec.id, {})
    return {
        "id": legacy.get("id", spec.id),
        "overseas_id": spec.id,
        "name": legacy.get("name", spec.name),
        "display_symbol": legacy.get("display_symbol", spec.display_symbol),
        "yahoo_symbol": spec.yahoo_symbol,
        "eastmoney_secid": spec.eastmoney_secid,
        "signal_symbol": spec.signal_symbol,
        "note": legacy.get("note", spec.note),
    }
